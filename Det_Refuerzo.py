# Calculo long de gancho estribos y barras
from Ecuaciones import*
import math

def Calcular():
    fc = float (c1ent02.get())
    fy = float (c1ent03.get())
    we = float (c1ent05.get())         
    wt = float (c1ent06.get())
    n = float (c1ent04.get())
    Nb = float (c1ent01.get())
    
    dbp = f_dbp(Nb);                       print(' dbp = ', dbp)
    c2ent01.delete(0, 'end');              c2ent01.insert (0, dbp)

    db = f_db(Nb);                         print(" db  = %.2f _ cm" %db)
    c2ent02.delete(0, 'end');              c2ent02.insert (0, "{:.2f}".format(db))

    Asb = f_Asb(Nb);                       print(" Asb = %.2f _ cm²" %Asb)
    c2ent03.delete(0, 'end');              c2ent03.insert (0, "{:.2f}".format(Asb))

    Wb = f_Wb(Nb);                         print(" Wb  = %.2f _ kg/m" %Wb)
    c2ent04.delete(0, 'end');              c2ent04.insert (0, "{:.2f}".format(Wb))

    # Gancho estribo a 90°
    print ("")
    print ("Long de gancho en estribos a 90° _ cm")

    if (Nb <= 5):
        D = 4*db
        K = max(7.5, 6*db)
        L = (1/4*math.pi*D)+K
        H = K + D/2 + db 
    else:
        D = 6*db
        K = 12*db
        L = (1/4*math.pi*D)+K
        H = K + D/2 + db
    print(' D = ', D);  c4ent04.delete(0, 'end');   c4ent04.insert (0, "{:.2f}".format(D))
    print(' K = ', K);  c4ent05.delete(0, 'end');   c4ent05.insert (0, "{:.2f}".format(K))
    print(' L = ', L);  c4ent06.delete(0, 'end');   c4ent06.insert (0, "{:.2f}".format(L))
    print(' H = ', H);  c4ent07.delete(0, 'end');   c4ent07.insert (0, "{:.2f}".format(H))

    # Gancho estribo a 135°
    print ("")
    print ("Long de gancho en estribos a 135° _ cm")
    K = max(7.5, 6*db)
    if (Nb <= 5):
        D = 4*db
        H = (0.707107*K)+D+db
        L = (0.375*math.pi*D)+K
    else:
        D = 6*db
        H = (0.707107*K)+D+db
        L = (0.375*math.pi*D)+K
    print(' D = ', D);  c4ent14.delete(0, 'end');   c4ent14.insert (0, "{:.2f}".format(D))
    print(' K = ', K);  c4ent15.delete(0, 'end');   c4ent15.insert (0, "{:.2f}".format(K))
    print(' L = ', L);  c4ent16.delete(0, 'end');   c4ent16.insert (0, "{:.2f}".format(L))
    print(' H = ', H);  c4ent17.delete(0, 'end');   c4ent17.insert (0, "{:.2f}".format(H))

    # Gancho estribo a 180°
    print ("")
    print ("Long de gancho en estribos a 180° _ cm")
    K = max(6.5, 4*db)
    if (Nb <= 5):
        D = 4*db
        L = (0.5*math.pi*D)+K
        H = D + 2*db 
    else:
        D = 6*db
        L = (0.5*math.pi*D)+K
        H = D + 2*db
    print(' D = ', D);  c4ent24.delete(0, 'end');   c4ent24.insert (0, "{:.2f}".format(D))
    print(' K = ', K);  c4ent25.delete(0, 'end');   c4ent25.insert (0, "{:.2f}".format(K))
    print(' L = ', L);  c4ent26.delete(0, 'end');   c4ent26.insert (0, "{:.2f}".format(L))
    print(' H = ', H);  c4ent27.delete(0, 'end');   c4ent27.insert (0, "{:.2f}".format(H))

    # Gancho en barras a 90°
    print ("")
    print ("Long de gancho en barras a 90° _ cm")

    if (Nb <= 8):
        D = 6*db
        K = 12*db
        L = (1/4*math.pi*D)+K
        H = K + D/2 + db 
    else:
        D = 8*db
        K = 8*db
        L = (1/4*math.pi*D)+K
        H = K + D/2 + db
    print(' D = ', D);  c5ent04.delete(0, 'end');   c5ent04.insert (0, "{:.2f}".format(D))
    print(' K = ', K);  c5ent05.delete(0, 'end');   c5ent05.insert (0, "{:.2f}".format(K))
    print(' L = ', L);  c5ent06.delete(0, 'end');   c5ent06.insert (0, "{:.2f}".format(L))
    print(' H = ', H);  c5ent07.delete(0, 'end');   c5ent07.insert (0, "{:.2f}".format(H))

    # Gancho barras a 180°
    print ("")
    print ("Long de gancho en barras a 180° _ cm")
    K = max(6.5, 4*db)
    if (Nb <= 8):
        D = 6*db
        L = (0.5*math.pi*D)+K
        H = D + 2*db 
    else:
        D = 8*db
        L = (0.5*math.pi*D)+K
        H = D + 2*db
    print(' D = ', D);  c5ent14.delete(0, 'end');   c5ent14.insert (0, "{:.2f}".format(D))
    print(' K = ', K);  c5ent15.delete(0, 'end');   c5ent15.insert (0, "{:.2f}".format(K))
    print(' L = ', L);  c5ent16.delete(0, 'end');   c5ent16.insert (0, "{:.2f}".format(L))
    print(' H = ', H);  c5ent17.delete(0, 'end');   c5ent17.insert (0, "{:.2f}".format(H))

    
    #Longitud de desarrollo ldh
    print ("")
    print ("Long de desarrollo ldh _ cm")

    ldh = max(0.8*db , (0.24*we*fy)/(n*math.sqrt(fc))*db);  print(' ldh = ', ldh)
    c3ent01.delete(0, 'end');   c3ent01.insert (0, "{:.2f}".format(ldh))

    ldh1 = 0.8*ldh;  print(' ldh1 = ', ldh1)
    c3ent02.delete(0, 'end');   c3ent02.insert (0, "{:.2f}".format(ldh1))
       
    # Longitud de desarrollo ld
    print ("")
    print ("Long de desarrollo en tracción ld _ cm")
    if (Nb <= 6):
        ld = (wt*we*fy)/(2.1*n*math.sqrt(fc))*db
    else:
        ld = (wt*we*fy)/(1.7*n*math.sqrt(fc))*db
    
    ldA = ld;           print(' ld clase A= ', ldA)
    c3ent03.delete(0, 'end');   c3ent03.insert (0, "{:.2f}".format(ldA))

    ldB = 1.3*ld;       print(' ld clase B= ', ldB)
    c3ent04.delete(0, 'end');   c3ent04.insert (0, "{:.2f}".format(ldB))
 

def Guardar():
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    book = Workbook()
    book = load_workbook('D:\IEB_REFUERZO\PlantillaRefuerzo.xlsx')
    sheet = book.active
    
    fc = float (c1ent02.get());     sheet['H4'] = "%.0f" %fc
    fy = float (c1ent03.get());     sheet['H5'] = "%.0f" %fy
    we = float (c1ent05.get());     sheet['H7'] = "%.1f" %we         
    wt = float (c1ent06.get());     sheet['H8'] = "%.1f" %wt
    n = float (c1ent04.get());      sheet['H6'] = "%.1f" %n
    Nb = float (c1ent01.get())     
    
    dbp = f_dbp(Nb);     sheet['D4'] = dbp
    db = f_db(Nb);       sheet['D5'] = "%.2f" %db                        
    Asb = f_Asb(Nb);     sheet['D6'] = "%.2f" %Asb                  
    Wb = f_Wb(Nb);       sheet['D7'] = "%.2f" %Wb 

    # Gancho estribo a 90°
    if (Nb <= 5):
        D = 4*db
        K = max(7.5, 6*db)
        L = (1/4*math.pi*D)+K
        H = K + D/2 + db 
    else:
        D = 6*db
        K = 12*db
        L = (1/4*math.pi*D)+K
        H = K + D/2 + db
    sheet['D12'] = "%.2f" %D
    sheet['D13'] = "%.2f" %K
    sheet['D14'] = "%.2f" %L
    sheet['D15'] = "%.2f" %H

    # Gancho estribo a 135°
    K = max(7.5, 6*db)
    if (Nb <= 5):
        D = 4*db
        H = (0.707107*K)+D+db
        L = (0.375*math.pi*D)+K
    else:
        D = 6*db
        H = (0.707107*K)+D+db
        L = (0.375*math.pi*D)+K
    sheet['E12'] = "%.2f" %D
    sheet['E13'] = "%.2f" %K
    sheet['E14'] = "%.2f" %L
    sheet['E15'] = "%.2f" %H

    # Gancho estribo a 180°
    K = max(6.5, 4*db)
    if (Nb <= 5):
        D = 4*db
        L = (0.5*math.pi*D)+K
        H = D + 2*db 
    else:
        D = 6*db
        L = (0.5*math.pi*D)+K
        H = D + 2*db
    sheet['F12'] = "%.2f" %D
    sheet['F13'] = "%.2f" %K
    sheet['F14'] = "%.2f" %L
    sheet['F15'] = "%.2f" %H

    # Gancho en barras a 90°
    if (Nb <= 8):
        D = 6*db
        K = 12*db
        L = (1/4*math.pi*D)+K
        H = K + D/2 + db 
    else:
        D = 8*db
        K = 8*db
        L = (1/4*math.pi*D)+K
        H = K + D/2 + db
    sheet['G12'] = "%.2f" %D
    sheet['G13'] = "%.2f" %K
    sheet['G14'] = "%.2f" %L
    sheet['G15'] = "%.2f" %H

    # Gancho barras a 180°
    K = max(6.5, 4*db)
    if (Nb <= 8):
        D = 6*db
        L = (0.5*math.pi*D)+K
        H = D + 2*db 
    else:
        D = 8*db
        L = (0.5*math.pi*D)+K
        H = D + 2*db
    sheet['H12'] = "%.2f" %D
    sheet['H13'] = "%.2f" %K
    sheet['H14'] = "%.2f" %L
    sheet['H15'] = "%.2f" %H

    #Longitud de desarrollo ldh
    ldh = max(0.8*db , (0.24*we*fy)/(n*math.sqrt(fc))*db); sheet['D18'] = "%.2f" %ldh  
    ldh1 = 0.8*ldh; sheet['D19'] = "%.2f" %ldh1 

    # Longitud de desarrollo ld
    if (Nb <= 6):
        ld = (wt*we*fy)/(2.1*n*math.sqrt(fc))*db
    else:
        ld = (wt*we*fy)/(1.7*n*math.sqrt(fc))*db
    
    ldA = ld;       sheet['H18'] = "%.2f" %ldA           
    ldB = 1.3*ld;   sheet['H19'] = "%.2f" %ldB
    
    print ("")
    print ("Los datos fueron guardados con exito")

    book.save ('D:\IEB_REPORTES\Reinforcement Details.xlsx')

def Borrar ():
    c1ent01.delete(0, 'end')
    c1ent02.delete(0, 'end')
    c1ent03.delete(0, 'end')
    c1ent04.delete(0, 'end')
    c1ent05.delete(0, 'end')
    c1ent06.delete(0, 'end')

    c2ent01.delete(0, 'end')
    c2ent02.delete(0, 'end')
    c2ent03.delete(0, 'end')
    c2ent04.delete(0, 'end')

    c3ent01.delete(0, 'end')
    c3ent02.delete(0, 'end')
    c3ent03.delete(0, 'end')
    c3ent04.delete(0, 'end')
    
    c4ent04.delete(0, 'end');       c4ent14.delete(0, 'end');       c4ent24.delete(0, 'end')
    c4ent05.delete(0, 'end');       c4ent15.delete(0, 'end');       c4ent25.delete(0, 'end')
    c4ent06.delete(0, 'end');       c4ent16.delete(0, 'end');       c4ent26.delete(0, 'end')
    c4ent07.delete(0, 'end');       c4ent17.delete(0, 'end');       c4ent27.delete(0, 'end')
    
    c5ent04.delete(0, 'end');       c5ent14.delete(0, 'end')
    c5ent05.delete(0, 'end');       c5ent15.delete(0, 'end')
    c5ent06.delete(0, 'end');       c5ent16.delete(0, 'end')
    c5ent07.delete(0, 'end');       c5ent17.delete(0, 'end')


# VENTANA DEL ESCRITORIO #
from tkinter import *
vent = Tk()
vent.geometry("620x760")
vent.title(" REINFORCEMENT DETAILS")
vent.iconbitmap('D:\\BIBLIOTECA PERSONAL\\Programación\\Python\\logo-wat.ico') 

# Recuadros

rec1 = LabelFrame(vent, text = ' Datos iniciales. ')
rec1.pack()
rec1.place(x=5, y=5, width=400, height=310)

rec2 = LabelFrame(vent, text = ' Propiedades de la barra ')
rec2.pack()
rec2.place(x=415, y=5, width=200, height=150)

rec3 = LabelFrame(vent, text = ' Longitud de desarrollo ldh y ld ')
rec3.pack()
rec3.place(x=415, y=165, width=200, height=150)

rec4 = LabelFrame(vent, text = ' Geometría gancho estandar - Estribos ')
rec4.pack()
rec4.place(x=5, y=325, width=400, height=170)

rec5 = LabelFrame(vent, text = ' Geometría gancho estandar - Barras ')
rec5.pack()
rec5.place(x=5, y=505, width=400, height=170)

rec6 = LabelFrame(vent, text = ' Diagramas ')
rec6.pack()
rec6.place(x=415, y=325, width=200, height=350)

rec7 = LabelFrame(vent, text = ' Opciones ')
rec7.pack()
rec7.place(x=110, y=675, width=400, height=55)

# Datos cuadro 1
c1tex01 = Label(rec1, text = "N° barra a usar"); c1tex01.pack()
c1tex01.place(x=5, y=10, width=280, height=20)
c1ent01 = Entry(rec1, justify=CENTER);                              c1ent01.place(x=295, y=10, width=90, height=20)

c1tex02 = Label(rec1, text = "fc: Resistencia del concreto _ MPa"); c1tex02.pack()
c1tex02.place(x=5, y=40, width=280, height=20)
c1ent02 = Entry(rec1, justify=CENTER);                              c1ent02.place(x=295, y=40, width=90, height=20)

c1tex03 = Label(rec1, text = "fy: Fluencia del acero _ MPa");       c1tex03.pack()
c1tex03.place(x=5, y=70, width=280, height=20)
c1ent03 = Entry(rec1, justify=CENTER);                              c1ent03.place(x=295, y=70, width=90, height=20)

c1tex04 = Label(rec1, text = "λ. Factor de modificación peso del concreto ¹"); c1tex04.pack()
c1tex04.place(x=5, y=100, width=280, height=20)
c1ent04 = Entry(rec1, justify=CENTER);                              c1ent04.place(x=295, y=100, width=90, height=20)

c1tex05 = Label(rec1, text = "Ψe. Factor de modificación por epóxico ¹"); c1tex05.pack()
c1tex05.place(x=5, y=130, width=280, height=20)
c1ent05 = Entry(rec1, justify=CENTER);                              c1ent05.place(x=295, y=130, width=90, height=20)

c1tex06 = Label(rec1, text = "Ψt. Factor de modificación por ubicación ¹"); c1tex06.pack()
c1tex06.place(x=5, y=160, width=280, height=20)
c1ent06 = Entry(rec1, justify=CENTER);                              c1ent06.place(x=295, y=160, width=90, height=20)

label = Label(rec1, text = "¹ ACI 318 - 19. Tabla 24.4.2.5", font='Arial 8', justify="left"); label.pack()
label.place(x=10, y=210, width=380, height=15)

label2 = Label(rec1, text = "² Longitud de desarrollo si se colocan estribos paralelos o ", font='Arial 8', justify="right")
label2.pack()
label2.place(x=10, y=230, width=380, height=15)

label3 = Label(rec1, text = "perpendiculares a la barra con gancho estándar", font='Arial 8', justify="left")
label3.pack()
label3.place(x=10, y=250, width=380, height=15)

label4 = Label(rec1, text = "³ - Empalme Clase A = 1.0 * ld.   - Empalme Clase B = 1.3 * ld", font='Arial 8', justify="left")
label4.pack()
label4.place(x=10, y=270, width=380, height=15)

# Datos cuadro 2
c2tex01 = Label(rec2, text = "ø: Diámetro _ in"); c2tex01.pack()
c2tex01.place(x=5, y=10, width=130, height=20)
c2ent01 = Entry(rec2, justify=CENTER);                              c2ent01.place(x=135, y=10, width=50, height=20); c2ent01.config(bg="#ecf0f1") 

c2tex02 = Label(rec2, text = "ø: Diámetro _ cm"); c2tex02.pack()
c2tex02.place(x=5, y=40, width=130, height=20)
c2ent02 = Entry(rec2, justify=CENTER);                              c2ent02.place(x=135, y=40, width=50, height=20); c2ent02.config(bg="#ecf0f1")

c2tex03 = Label(rec2, text = "Área _ cm²"); c2tex03.pack()
c2tex03.place(x=5, y=70, width=130, height=20)
c2ent03 = Entry(rec2, justify=CENTER);                              c2ent03.place(x=135, y=70, width=50, height=20); c2ent03.config(bg="#ecf0f1") 

c2tex04 = Label(rec2, text = "Peso unitario _ kg/m"); c2tex04.pack()
c2tex04.place(x=5, y=100, width=130, height=20)
c2ent04 = Entry(rec2, justify=CENTER);                              c2ent04.place(x=135, y=100, width=50, height=20); c2ent04.config(bg="#ecf0f1")

# Datos cuadro 3
c3tex01 = Label(rec3, text = "ldh _ cm"); c3tex01.pack()
c3tex01.place(x=5, y=10, width=130, height=20)
c3ent01 = Entry(rec3, justify=CENTER);                              c3ent01.place(x=135, y=10, width=50, height=20); c3ent01.config(bg="#ecf0f1") 

c3tex02 = Label(rec3, text = "ldh1 ² _ cm"); c3tex02.pack()
c3tex02.place(x=5, y=40, width=130, height=20)
c3ent02 = Entry(rec3, justify=CENTER);                              c3ent02.place(x=135, y=40, width=50, height=20); c3ent02.config(bg="#ecf0f1")

c3tex03 = Label(rec3, text = "ld clase A ³ _ cm"); c3tex03.pack()
c3tex03.place(x=5, y=70, width=130, height=20)
c3ent03 = Entry(rec3, justify=CENTER);                              c3ent03.place(x=135, y=70, width=50, height=20); c3ent03.config(bg="#ecf0f1") 

c3tex04 = Label(rec3, text = "ld clase B ³ _ cm"); c3tex04.pack()
c3tex04.place(x=5, y=100, width=130, height=20)
c3ent04 = Entry(rec3, justify=CENTER);                              c3ent04.place(x=135, y=100, width=50, height=20); c3ent04.config(bg="#ecf0f1")

# Datos cuadro 4
c4tex01 = Label(rec4, text = "Gancho a 90°", font='Helvetica 7 bold'); c4tex01.pack()
c4tex01.place(x=160, y=5, width=70, height=20)

c4tex02 = Label(rec4, text = "Gancho a 135°", font='Helvetica 7 bold'); c4tex02.pack()
c4tex02.place(x=240, y=5, width=70, height=20)

c4tex03 = Label(rec4, text = "Gancho a 180°", font='Helvetica 7 bold'); c4tex03.pack()
c4tex03.place(x=320, y=5, width=70, height=20)

c4tex04 = Label(rec4, text = "D: ø de doblado _cm");                c4tex04.pack();     c4tex04.place(x=5, y=30, width=140, height=20)
c4ent04 = Entry(rec4, justify=CENTER);                              c4ent04.place(x=160, y=30, width=70, height=20); c4ent04.config(bg="#ecf0f1")
c4ent14 = Entry(rec4, justify=CENTER);                              c4ent14.place(x=240, y=30, width=70, height=20); c4ent14.config(bg="#ecf0f1")
c4ent24 = Entry(rec4, justify=CENTER);                              c4ent24.place(x=320, y=30, width=70, height=20); c4ent24.config(bg="#ecf0f1")

c4tex05 = Label(rec4, text = "K: Extención recta _cm");             c4tex05.pack();     c4tex05.place(x=5, y=60, width=140, height=20)
c4ent05 = Entry(rec4, justify=CENTER);                              c4ent05.place(x=160, y=60, width=70, height=20); c4ent05.config(bg="#ecf0f1")
c4ent15 = Entry(rec4, justify=CENTER);                              c4ent15.place(x=240, y=60, width=70, height=20); c4ent15.config(bg="#ecf0f1")
c4ent25 = Entry(rec4, justify=CENTER);                              c4ent25.place(x=320, y=60, width=70, height=20); c4ent25.config(bg="#ecf0f1")

c4tex06 = Label(rec4, text = "L: Long gancho _cm");                 c4tex06.pack();     c4tex06.place(x=5, y=90, width=140, height=20)
c4ent06 = Entry(rec4, justify=CENTER);                              c4ent06.place(x=160, y=90, width=70, height=20); c4ent06.config(bg="#ecf0f1")
c4ent16 = Entry(rec4, justify=CENTER);                              c4ent16.place(x=240, y=90, width=70, height=20); c4ent16.config(bg="#ecf0f1")
c4ent26 = Entry(rec4, justify=CENTER);                              c4ent26.place(x=320, y=90, width=70, height=20); c4ent26.config(bg="#ecf0f1")

c4tex07 = Label(rec4, text = "H: Altura _cm");                      c4tex07.pack();     c4tex07.place(x=5, y=120, width=140, height=20)
c4ent07 = Entry(rec4, justify=CENTER);                              c4ent07.place(x=160, y=120, width=70, height=20); c4ent07.config(bg="#ecf0f1")
c4ent17 = Entry(rec4, justify=CENTER);                              c4ent17.place(x=240, y=120, width=70, height=20); c4ent17.config(bg="#ecf0f1")
c4ent27 = Entry(rec4, justify=CENTER);                              c4ent27.place(x=320, y=120, width=70, height=20); c4ent27.config(bg="#ecf0f1")

# Datos cuadro 5

c5tex01 = Label(rec5, text = "Gancho a 90°", font='Helvetica 7 bold'); c5tex01.pack()
c5tex01.place(x=200, y=5, width=70, height=20)

c5tex02 = Label(rec5, text = "Gancho a 180°", font='Helvetica 7 bold'); c5tex02.pack()
c5tex02.place(x=280, y=5, width=70, height=20)

c5tex04 = Label(rec5, text = "D: ø de doblado _cm");                c5tex04.pack();     c5tex04.place(x=50, y=30, width=140, height=20)
c5ent04 = Entry(rec5, justify=CENTER);                              c5ent04.place(x=200, y=30, width=70, height=20); c5ent04.config(bg="#ecf0f1")
c5ent14 = Entry(rec5, justify=CENTER);                              c5ent14.place(x=280, y=30, width=70, height=20); c5ent14.config(bg="#ecf0f1")

c5tex05 = Label(rec5, text = "K: Extención recta _cm");             c5tex05.pack();     c5tex05.place(x=50, y=60, width=140, height=20)
c5ent05 = Entry(rec5, justify=CENTER);                              c5ent05.place(x=200, y=60, width=70, height=20); c5ent05.config(bg="#ecf0f1")
c5ent15 = Entry(rec5, justify=CENTER);                              c5ent15.place(x=280, y=60, width=70, height=20); c5ent15.config(bg="#ecf0f1")

c5tex06 = Label(rec5, text = "L: Long gancho _cm");                 c5tex06.pack();     c5tex06.place(x=50, y=90, width=140, height=20)
c5ent06 = Entry(rec5, justify=CENTER);                              c5ent06.place(x=200, y=90, width=70, height=20); c5ent06.config(bg="#ecf0f1")
c5ent16 = Entry(rec5, justify=CENTER);                              c5ent16.place(x=280, y=90, width=70, height=20); c5ent16.config(bg="#ecf0f1")

c5tex07 = Label(rec5, text = "H: Altura _cm");                      c5tex07.pack();     c5tex07.place(x=50, y=120, width=140, height=20)
c5ent07 = Entry(rec5, justify=CENTER);                              c5ent07.place(x=200, y=120, width=70, height=20); c5ent07.config(bg="#ecf0f1")
c5ent17 = Entry(rec5, justify=CENTER);                              c5ent17.place(x=280, y=120, width=70, height=20); c5ent17.config(bg="#ecf0f1")

# Datos cuadro 6 imagen

from PIL import ImageTk, Image
img = Image.open('D:\IEB_REFUERZO\Diagramas.jpg')
img = img.resize((170,340), Image.ANTIALIAS)

img = ImageTk.PhotoImage(img)
lbl_img = Label(rec6, image=img)
lbl_img.pack()

# Recuadro 7 botones

bot1 = Button(rec7, text = 'Calcular', font='Helvetica 8 bold', command=Calcular); bot1.pack()
bot1.place(x=40, y=5, width=80, height=20)

bot2 = Button(rec7, text = 'Guardar .xls', font='Helvetica 8 bold', command=Guardar); bot2.pack()
bot2.place(x=160, y=5, width=80, height=20)

bot3 = Button(rec7, text = 'Borrar', font='Helvetica 8 bold', command=Borrar); bot3.pack()
bot3.place(x=280, y=5, width=80, height=20)

# firma
label = Label(vent, text = "wilson.taimalc@gmail.com - 2023", font='Arial 7'); label.pack()
label.place(x=10, y=740, width=600, height=10)

vent.mainloop()




